#!/usr/bin/env python3
"""
docs/14-PRODUCT-INVENTORY.xlsx — every surface, every job, every command, with a verdict.

GENERATED, NEVER TYPED. It reads the sources of truth so the workbook cannot drift from the
product it describes:

  · `chute finder-actions --json`     the 14 actions, live from the binary
  · site/src/lib/cases.ts             25 cases: the pain, the ritual, the fix, the arithmetic
  · docs/03-JTBD-LEDGER.md            the 24 jobs
  · docs/05-CUSTOMER-JOURNEY-MAP.md   the six stages
  · `chute help`                      the command list

The one sheet that is written by hand is `Menu Bar`, because a verdict is a judgement and there is
nowhere to read one from. Everything on it names the file and line its evidence came from.

  python3 Scripts/build-inventory-xlsx.py
"""
import json, os, re, subprocess, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "docs", "14-PRODUCT-INVENTORY.xlsx")

INK, GROUND, ACCENT, MUTED = "FFF7F7F7", "FF171C29", "FF8FDB70", "FF8A93A6"
HEAD = PatternFill("solid", fgColor=GROUND)
ZEBRA = PatternFill("solid", fgColor="FFF3F5F8")
THIN = Border(bottom=Side(style="thin", color="FFD9DEE7"))


def sh(cmd, cwd=ROOT):
    p = subprocess.run(cmd, shell=True, cwd=cwd, capture_output=True, text=True)
    return p.stdout


def cases():
    out = sh("node --input-type=module -e \"import {CASES} from './src/lib/cases.ts';"
             "console.log(JSON.stringify(CASES))\" 2>/dev/null", cwd=os.path.join(ROOT, "site"))
    return json.loads(out.strip().split("\n")[-1])


def actions():
    return json.loads(sh(".build/release/chute finder-actions --json"))


def ledger():
    rows = {}
    for line in open(os.path.join(ROOT, "docs", "03-JTBD-LEDGER.md")):
        cell = [c.strip() for c in line.split("|")]
        if len(cell) < 8:
            continue
        try:
            n = int(cell[1])
        except ValueError:
            continue
        num = lambda s: (lambda m: float(m[0]) if m else None)(re.findall(r"-?[\d.]+", s))
        rows[n] = dict(name=cell[2], perDay=num(cell[3]), manual=num(cell[4]),
                       chute=num(cell[5]), saved=num(cell[6]), cost=cell[7],
                       tier=cell[8] if len(cell) > 8 else "")
    return rows


def journey():
    rows, path = [], os.path.join(ROOT, "docs", "05-CUSTOMER-JOURNEY-MAP.md")
    for line in open(path):
        cell = [c.strip() for c in line.split("|")]
        if len(cell) >= 8 and cell[1] and not set(cell[1]) <= set("-: ") and cell[1] != "Stage":
            rows.append(cell[1:9])
    return rows


def commands():
    rows, section = [], ""
    for line in sh(".build/release/chute help").split("\n"):
        if re.match(r"^[A-Z][A-Z ]+$", line.strip()):
            section = line.strip()
        # Columns in `chute help` are separated by runs of two or more spaces, and a command's
        # signature contains single spaces ("paths <files…>", "buf add|list|flush|clear"). Matching
        # \S+ for the name stopped at "paths" and then needed two spaces where there was one, so
        # twenty-one of the twenty-seven commands silently did not appear in the workbook.
        if not line.startswith("  ") or line.startswith("   ") or not line.strip():
            continue
        parts = [p for p in re.split(r"\s{2,}", line.strip()) if p]
        if len(parts) < 2 or parts[0].startswith("-"):
            continue
        name = parts[0].split(" ")[0].split("|")[0]
        flags = " ".join(p for p in parts[1:] if p.startswith("-"))
        purpose = " ".join(p for p in parts[1:] if not p.startswith("-"))
        rows.append([section, name, purpose, flags])
    return rows


def sheet(wb, title, headers, rows, widths):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color=INK, name="Menlo", size=10)
        c.fill = HEAD
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    for i, row in enumerate(rows):
        ws.append(row)
        for c in ws[ws.max_row]:
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.font = Font(size=10)
            c.border = THIN
            if i % 2:
                c.fill = ZEBRA
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return ws


def main():
    if not os.path.exists(os.path.join(ROOT, ".build/release/chute")):
        sys.exit("build first: swift build -c release")
    CS, ACT, LG = cases(), actions(), ledger()
    by_jtbd = {c["jtbd"]: c for c in CS if c.get("jtbd")}
    by_cmd = {c.get("command", "").replace("chute ", "").split(" ")[0]: c for c in CS}

    wb = Workbook()
    wb.remove(wb.active)

    # ── 1. Finder menu ──────────────────────────────────────────────────────────────────────
    rows = []
    for a in ACT:
        verb = a["argv"][0]
        c = by_cmd.get(verb, {})
        j = LG.get(c.get("jtbd"), {})
        rows.append([
            a["id"], a["title"], "submenu" if a["parentTitle"] else "top-level row",
            a["parentTitle"] or "—", a["scope"],
            "confirms: " + a["confirmButton"] if a["confirmButton"] else "no confirmation",
            "chute " + " ".join(a["argv"]), a["detail"],
            c.get("jtbd", "—"), j.get("perDay", "—"), j.get("manual", "—"), j.get("chute", "—"),
            j.get("saved", "—"), c.get("pain", "—"), c.get("ritual", "—"), c.get("fix", "—"),
        ])
    sheet(wb, "Finder Menu",
          ["Action id", "Menu title", "Drawn as", "Submenu", "Acts on", "Safety", "CLI it runs",
           "What it does", "JTBD", "Times/day", "By hand (s)", "With Chute (s)", "Saved min/day",
           "The pain", "What you do now", "What happens instead"],
          rows, [20, 30, 14, 20, 11, 22, 30, 42, 7, 10, 12, 13, 13, 46, 46, 46])

    # ── 2. Menu bar — the verdict sheet ─────────────────────────────────────────────────────
    mb = [
        ["Menu bar extra (⤓ + count)", "Template NSImage + the number of sessions wanting you",
         "Tell me, without looking, whether anything needs me", "0 — attention, not time", "glanced at constantly",
         "KEPT, CHANGED", "Was the literal text '⤓ 3'. A glyph drawn as text does not participate in the "
         "menu bar's tinting, so it came out wrong on a light bar and in reduced contrast. Every Apple "
         "extra is a template image. handoff/NEXT.md:295", "Template image, count beside it, nothing at zero"],
        ["Waiting for You / Working / Idle", "Sessions grouped by what you must DO about them",
         "Nine terminals, one stopped four minutes ago and I don't know which", "0", "many times a day",
         "KEPT, CHANGED", "Rendered as disabled items. NSMenuItem.sectionHeader(title:) is the HIG API and "
         "exists on macOS 14; the fallback stays for macOS 13. Idle now collapses past three — it is the "
         "group you never act on and it pushed the others off the top.", "Real section headers; idle collapses"],
        ["Session row", "project · agent · model · effort · waited · CPU · memory",
         "Which agent is this, what is it running on, and is it stuck", "0", "many times a day",
         "KEPT, CHANGED", "The row said 'Claude Code' — that was the terminal WINDOW TITLE leaking through, "
         "a string Chute had not derived. TerminalAppAdapter.swift:74 matched the agent's name and threw it "
         "away; cursor was missing entirely. Model and effort come from the agent's own transcript.",
         "Agent, model and effort established rather than echoed"],
        ["Session row → click", "Brings that terminal forward", "Get me to the one that is waiting",
         "0", "many times a day", "KEPT", "The hero job. It is why per-session commands are ⌥ alternates "
         "and not a submenu: an NSMenuItem with a submenu does not fire its own action.", "Unchanged"],
        ["Session row → ⌥", "Copy Session ID · Copy Resume Command · Continue in tmux · Copy Cost",
         "I need this conversation somewhere else / what has it cost me", "0", "occasional",
         "NEW", "The session id was parsed by HookRecord and emitted by nothing. tmux CONTINUES, never "
         "'moves': macOS cannot transplant a running process onto a new tty.", "Added"],
        ["1% CPU · 974 MB memory", "Per-session cost", "Which agent has gone wrong", "0", "glanced at",
         "KEPT — removed, then restored", "Removed as noise, restored on the owner's instruction: these are "
         "the useful part when you are comparing five agents. Now shown for EVERY session, because a blank "
         "reads as 'not measured' rather than 'small'.", "On every row, plus ⚠ above a threshold"],
        ["This Mac — using 0.4 of 16 cores · battery at 31 °C · 87 °F", "Whole-machine summary",
         "None", "—", "—", "DELETED", "The battery sensor is not the machine's temperature — it sits in the "
         "battery pack and does not track the chassis under an agent workload; the owner bought a cooling pad "
         "while this read 'running cool'. A number that disagrees with the hand on the keyboard teaches you to "
         "disbelieve the whole menu. The core average never explained why anything was slow.", "Gone"],
        ["Local Servers (n)", "Every listener, with the project it belongs to; open, copy, stop",
         "Something is on 3000 and I cannot find which window", "15", "11×/day",
         "KEPT", "Validated in the field 2026-08-28: the owner used it to find three idle services "
         "(postgres, redis, ollama) and removed them permanently.", "Unchanged"],
        ["Clipboard Buffer ▸ Add Clipboard to Buffer", "Manual capture of the clipboard",
         "The files I need are in four folders and the clipboard holds one", "22", "12×/day",
         "REPLACED", "A ritual: copy, then remember to open the menu and press Add — at the exact moment you "
         "are least likely to be thinking about it, and if you forget the thing is already gone.",
         "Recent Copies: everything Chute hands you is remembered as it is handed over"],
        ["Recent Copies (n)", "The last 10 things Chute copied for you; click one to copy it back",
         "I bundled three folders and the third overwrote the first", "22", "12×/day",
         "NEW", "Records what Chute WROTE, never what it read — no pasteboard observer, no changeCount poll. "
         "A password from a manager cannot appear in it, which is what lets it be on by default.",
         "Added"],
        ["Trial — 14 days left", "Days remaining, every single open", "None — it is a sales line",
         "—", "every open", "KEPT, CHANGED", "The code's own rule is that an app which keeps mentioning "
         "payment is nagging its customer. Fourteen reminders over fourteen days is fourteen chances to find "
         "the menu annoying, and the first eleven cannot change a decision nobody is making yet.",
         "Final 3 days only; Settings → License always shows it"],
        ["Refresh Now", "Rebuilt the menu", "None", "—", "—", "DELETED",
         "It could not have worked. refresh() built a NEW NSMenu and assigned it to statusItem.menu; "
         "menuWillOpen then fired on that fresh object and rebuilt everything from scratch, discarding it. "
         "Three automatic paths already kept the menu current. A command that cannot change what you see "
         "teaches the reader that the menu might be stale.", "Gone"],
        ["Report a Problem…", "Diagnostics to the clipboard, prefilled GitHub issue", "Support without an inbox",
         "—", "rare", "KEPT", "One public answer serves everyone who searches for the same thing.", "Unchanged"],
        ["Settings…", "General / License / About", "Activate a licence; see the real trial state",
         "—", "rare", "KEPT", "Correctly ellipsised: it opens a window needing more input.", "Unchanged"],
        ["Quit Chute", "Quits", "Quit", "—", "rare", "KEPT",
         "No ⌘Q shown: in a status menu a key equivalent works only while the menu is open, so printing one "
         "promises a global shortcut that does not exist.", "Unchanged"],
    ]
    sheet(wb, "Menu Bar",
          ["Item", "What it does", "The job it serves", "JTBD", "How often", "Verdict",
           "Why — with the evidence", "After"],
          mb, [34, 40, 40, 8, 16, 20, 74, 40])

    # ── 3. CLI ──────────────────────────────────────────────────────────────────────────────
    finder_verbs = {a["argv"][0] for a in ACT}
    MENUBAR = {"ports", "sessions", "focus", "doctor", "resume", "buf"}
    rows = []
    for section, name, purpose, flags in commands():
        c = by_cmd.get(name, {})
        j = LG.get(c.get("jtbd"), {})
        rows.append([section.title(), name, purpose, flags,
                     c.get("jtbd", "—"), c.get("slug", "—"),
                     "yes" if name in finder_verbs else "no",
                     "yes" if name in MENUBAR else "no",
                     j.get("saved", "—"), "free (MIT)" if not c or not c.get("paid") else "app"])
    sheet(wb, "CLI",
          ["Group", "Command", "What it does", "Flags", "JTBD", "Case", "In Finder menu?",
           "In menu bar?", "Saved min/day", "Free or paid"],
          rows, [16, 14, 46, 34, 7, 34, 15, 14, 13, 13])

    # ── 4. JTBD ledger ──────────────────────────────────────────────────────────────────────
    rows = []
    for n in sorted(LG):
        j, c = LG[n], by_jtbd.get(n, {})
        surfaces = [s for s, ok in (("Finder", c.get("surface") == "finder"),
                                    ("Menu bar", c.get("surface") == "menubar"),
                                    ("CLI", bool(c.get("command"))))
                    if ok]
        rows.append([n, j["name"], j["perDay"], j["manual"], j["chute"], j["saved"], j["cost"],
                     j["tier"], ", ".join(surfaces) or "not surfaced",
                     c.get("slug", "no case"), "yes" if c.get("demo") else "no recording"])
    sheet(wb, "JTBD Ledger",
          ["#", "Job to be done", "Times/day", "By hand (s)", "With Chute (s)", "Saved min/day",
           "Build cost", "Tier", "Surfaces that cover it", "Case page", "Recorded?"],
          rows, [5, 40, 10, 12, 14, 13, 11, 7, 24, 38, 14])

    # ── 5. Cases and the journey they belong to ─────────────────────────────────────────────
    rows = [[c["slug"], c.get("jtbd", "—"), c["surface"], c["tier"],
             "paid app" if c["paid"] else "free CLI", c["pain"], c["ritual"], c["fix"],
             c["seconds"]["manual"], c["seconds"]["chute"], c["perDay"],
             c.get("savedMinutes") if c.get("savedMinutes") is not None else "no figure — on purpose",
             c.get("command", "—"), c.get("demo", "not recorded")] for c in CS]
    sheet(wb, "Cases",
          ["Slug", "JTBD", "Surface", "Tier", "Who pays", "The pain, in their words",
           "What you do now", "What happens instead", "By hand (s)", "With Chute (s)",
           "Times/day", "Saved min/day", "CLI equivalent", "Recording"],
          rows, [40, 7, 11, 8, 11, 50, 50, 50, 12, 14, 11, 13, 34, 34])

    # ── 6. Customer journey ─────────────────────────────────────────────────────────────────
    sheet(wb, "Journey",
          ["Stage", "Doing", "Thinking", "Feeling", "Friction to remove", "Our move",
           "Success signal", "Which surface serves it"],
          journey(), [16, 40, 34, 16, 44, 44, 30, 30])

    # ── 7. Proposed, and refused ────────────────────────────────────────────────────────────
    prop = [
        ["Agent, model and effort on every row", "I have Opus in one window and Sonnet in another and "
         "cannot tell which is which", "0", "many/day", "S", "BUILT",
         "Reads the agent's own transcript, read-only. Every field optional: a format change costs a "
         "detail, never a crash.", "Nothing leaves the machine; ~/.claude is never written to"],
        ["Copy Session ID / Copy Resume Command", "I want this conversation in another window",
         "0", "occasional", "S", "BUILT", "The id was already parsed and never emitted.",
         "The id is already on disk in a filename"],
        ["Continue in tmux", "I need this to survive closing the lid", "0", "occasional", "S", "BUILT",
         "RE-CREATES the session; macOS cannot transplant a running process onto a new tty. The command "
         "is copied, never run.", "Copied, not executed — nothing is launched for you"],
        ["Session cost so far", "What has this session actually cost me", "0", "daily", "S", "BUILT",
         "Tokens, never dollars: a price is a number on someone else's pricing page.", "Local file only"],
        ["Recent Copies", "I bundled three folders and the third overwrote the first", "22", "12/day",
         "S", "BUILT", "Records what Chute wrote, never what it read.",
         "No pasteboard observer — a password from a manager cannot appear"],
        ["Passive clipboard history", "Same job, captured automatically from anything you copy",
         "22", "12/day", "M", "REFUSED", "A background pasteboard watcher is a permission story, a trust "
         "story and a support burden. Recent Copies does the job with none of it.",
         "Would silently hold passwords, keys and private messages"],
        ["Moving a live process into tmux", "Move this running agent to tmux", "—", "—", "—", "IMPOSSIBLE",
         "reptyr depends on Linux ptrace semantics; there is no macOS equivalent. Claiming it would be the "
         "first false thing this product said.", "—"],
        ["Dollar cost per session", "What am I spending", "—", "daily", "S", "DEFERRED",
         "A hardcoded price is wrong the week the vendor changes it. Needs a rate table with a retrievedOn "
         "date that the UI quotes.", "—"],
        ["Transcript readers for Codex / Cursor / Gemini", "Same detail for my other agents", "—", "—",
         "M each", "DEFERRED", "They already get their NAME free. A reader each is four vendor couplings "
         "for an audience of unknown size — build the second when someone asks.", "—"],
        ["A ninth Finder row", "Surface diff / redact / latest / gist / dataurl from Finder", "—", "—",
         "S each", "REFUSED", "docs/12-CAPABILITY-MAP.md §C: eight rows is the budget. The six deferred jobs "
         "are worth ~23 min/day combined — less than Copy Files as Context returns alone.", "—"],
    ]
    sheet(wb, "Proposed",
          ["Capability", "The pain, in their words", "JTBD", "How often", "Build cost", "Decision",
           "Reasoning", "Privacy note"],
          prop, [40, 52, 7, 13, 12, 14, 62, 46])

    wb.save(OUT)
    print(f"  {OUT}")
    for ws in wb:
        print(f"    {ws.title:<14} {ws.max_row - 1:>3} rows × {ws.max_column} cols")


if __name__ == "__main__":
    main()
